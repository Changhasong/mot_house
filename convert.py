#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
"""
맛집 데이터 변환 스크립트
맛집데이터.xlsx (또는 맛집데이터.csv) → data/restaurants.js

사용법: 데이터변환.bat 더블클릭  (또는  python convert.py)

엑셀 열 순서:
  A: 이름
  B: 카테고리        korean / chinese / japanese / cafe / bar / lunch
                     여러 카테고리는 쉼표로 구분: korean,lunch
  C: 설명            카드에 표시할 짧은 소개
  D: 위치(짧게)      카드에 표시할 짧은 주소
  E: 추천메뉴        쉼표로 구분
  F: 별점            1~5 (0.5 단위)
  G: 이미지파일      images/ 폴더 안의 파일명 (없으면 빈칸)
  H: 주소(전체)      상세 페이지에 표시할 전체 주소
  I: 영업시간        여러 줄은 셀 안에서 Alt+Enter로 줄바꿈
  J: 가격대          예: 8,000 – 12,000원
  K: 주차
  L: 전화번호
  M: 방문일          예: 2025-03-15
  N: 후기            긴 텍스트. 문단 구분은 셀 안에서 빈 줄(Alt+Enter 두 번)
  O: 맛별점          1~5 (0.5 단위)
  P: 가성비별점      1~5 (0.5 단위)
  Q: 분위기별점      1~5 (0.5 단위)
  R: 방문팁
  S: 추천태그        쉼표로 구분: 가성비 찾는 분, 점심 혼밥러
  T: 추가사진        쉼표로 구분된 파일명 (images/ 폴더 기준)
  U: 지도URL         카카오맵 / 네이버지도 링크
"""

import os
import json
import csv

XLSX_FILE   = '맛집데이터.xlsx'
CSV_FILE    = '맛집데이터.csv'
OUTPUT_JSON = 'data/restaurants.json'

VALID_CATS = {'korean', 'chinese', 'japanese', 'cafe', 'bar', 'lunch'}

HEADERS = [
    '이름', '카테고리', '설명', '위치(짧게)', '추천메뉴', '별점', '이미지파일',
    '주소(전체)', '영업시간', '가격대', '주차', '전화번호',
    '방문일', '후기', '맛별점', '가성비별점', '분위기별점',
    '방문팁', '추천태그', '추가사진', '지도URL'
]

SAMPLE_ROWS = [
    [
        '용문 가마솥 순두부', 'korean',
        '구수한 가마솥 순두부와 돌솥밥이 일품인 동네 맛집.',
        '대전 서구 용문동',
        '순두부찌개, 돌솥비빔밥', 4.5, '',
        '대전 서구 용문동 [번지수 입력]',
        '평일 08:00 – 21:00\n토요일 08:00 – 20:00\n일요일 정기 휴무',
        '8,000 – 12,000원 (1인 기준)',
        '건물 앞 소규모 주차 가능 (3대)', '042-000-0000', '2025-01-01',
        '용문동에서 오래됐다는 소문을 듣고 드디어 방문했습니다.\n\n순두부찌개는 뚝배기에 보글보글 끓는 상태로 나왔는데 정말 맛있었습니다.',
        4.5, 5.0, 4.0,
        '점심시간(12:00–13:00)은 매우 혼잡합니다. 11:30 이전 방문 권장.',
        '가성비 찾는 분, 점심 혼밥러, 어르신 모시고 온 가족', '', ''
    ],
    [
        '향미 중화반점', 'chinese',
        '동네에서 30년 넘게 운영 중인 중화요리 노포.',
        '대전 서구 용문동',
        '짜장면, 탕수육, 짬뽕', 4.0, '',
        '대전 서구 용문동 [번지수 입력]',
        '매일 11:00 – 21:00\n브레이크타임 15:00 – 17:00\n월요일 정기 휴무',
        '6,000 – 25,000원',
        '인근 공영주차장 이용', '042-000-0000', '2025-01-01',
        '30년 넘게 한자리를 지킨 노포답게 음식 맛만큼은 확실합니다.\n\n춘장을 직접 볶아 만든 짜장소스가 깊고 구수한 맛을 냅니다.',
        4.0, 4.5, 3.5,
        '브레이크타임(15:00–17:00)을 꼭 확인하세요.',
        '30년 노포 팬, 짜장면 마니아, 가족 외식', '', ''
    ],
    [
        '스시 히나', 'japanese',
        '합리적인 가격의 스시 오마카세 코스. 예약 필수.',
        '대전 서구 용문동',
        '스시 오마카세, 연어초밥', 5.0, '',
        '대전 서구 용문동 [번지수 입력]',
        '화–일 12:00 – 22:00\n월요일 정기 휴무\n예약제 운영',
        '30,000 – 60,000원 (코스 기준)',
        '주차 불가 (대중교통 이용 권장)', '042-000-0000', '2025-01-01',
        '예약하고 가야 한다는 게 번거롭지만, 막상 방문하니 그 이유를 알 것 같았습니다.\n\n셰프가 직접 한 피스씩 쥐어주는 스시는 신선도와 밥의 온도가 완벽했습니다.',
        5.0, 4.5, 5.0,
        '반드시 전화 예약 필수. 주말은 1–2주 전 예약을 권장합니다.',
        '특별한 날 외식, 스시 마니아, 데이트 코스', '', ''
    ],
    [
        '카페 온더로드', 'cafe',
        '용문동 골목 안 아늑한 독립 카페. 직접 로스팅한 원두.',
        '대전 서구 용문동',
        '핸드드립 커피, 바스크 치즈케이크', 4.5, '',
        '대전 서구 용문동 [번지수 입력]',
        '매일 10:00 – 22:00\n연중무휴',
        '4,500 – 8,000원',
        '골목 내 주차 어려움 (도보 방문 권장)', '042-000-0000', '2025-01-01',
        '골목 안쪽에 숨어 있는 카페는 내부가 넓고 아늑했습니다.\n\n핸드드립 커피는 원두를 직접 설명해 주고 취향에 맞게 골라주셨습니다.',
        4.5, 4.0, 5.0,
        '주말 오후는 자리가 없을 수 있으니 평일 방문 추천. 노트북 사용 가능.',
        '카페 투어족, 작업하기 좋은 곳, 데이트 코스', '', ''
    ],
    [
        '용문 포차', 'bar',
        '퇴근 후 한잔하기 딱 좋은 동네 포장마차 스타일 술집.',
        '대전 서구 용문동',
        '닭발, 순대볶음, 막걸리', 4.0, '',
        '대전 서구 용문동 [번지수 입력]',
        '매일 17:00 – 01:00\n명절 당일 휴무',
        '안주 10,000 – 20,000원 / 술 3,000원~',
        '인근 골목 주차 가능', '042-000-0000', '2025-01-01',
        '퇴근길에 들렀는데 이미 자리가 거의 찼을 정도로 인기 있는 집입니다.\n\n닭발은 매콤달콤한 양념이 잘 배어 있고 양도 넉넉했습니다.',
        4.0, 4.5, 4.0,
        '금·토요일 저녁은 자리가 빨리 찹니다. 인원이 많으면 미리 전화 확인하세요.',
        '퇴근 후 한잔, 가성비 술집, 소그룹 모임', '', ''
    ],
    [
        '점심특선 한식당', 'korean,lunch',
        '직장인 점심 성지. 7,000원에 국밥 + 반찬 무한 제공.',
        '대전 서구 용문동',
        '점심특선 정식, 소고기국밥', 5.0, '',
        '대전 서구 용문동 [번지수 입력]',
        '평일 11:00 – 14:00\n주말·공휴일 휴무',
        '7,000 – 9,000원',
        '건물 공용주차장 이용 (30분 무료)', '042-000-0000', '2025-01-01',
        '11시 30분에 도착했는데 이미 자리가 반 이상 찼습니다.\n\n7,000원에 국밥과 반찬이 무한으로 나오는 구성이 믿기지 않을 정도입니다.',
        5.0, 5.0, 4.5,
        '11:30 이전에 가야 자리를 잡을 수 있습니다. 14:00 이후에는 문을 닫습니다.',
        '직장인 점심, 초가성비, 혼밥 가능', '', ''
    ],
]


def col(row, n, default=''):
    v = row[n] if len(row) > n and row[n] is not None else None
    if v is None:
        return default
    s = str(v).strip()
    return s if s and s != 'None' else default


def parse_rating(row, n, fallback=4.0):
    try:
        r = round(float(row[n]) * 2) / 2 if len(row) > n and row[n] else fallback
        return max(0.5, min(5.0, r))
    except (TypeError, ValueError):
        return fallback


def parse_row(row):
    name = col(row, 0)
    if not name:
        return None

    cat_raw  = col(row, 1, 'korean')
    desc     = col(row, 2)
    location = col(row, 3, '대전 서구 용문동')
    menu     = col(row, 4)
    rating   = parse_rating(row, 5, 4.0)
    _img_raw = col(row, 6)
    image    = (_img_raw if '.' in _img_raw else _img_raw + '.jpg') if _img_raw else ''
    address  = col(row, 7) or location
    hours    = col(row, 8)
    price    = col(row, 9)
    parking  = col(row, 10)
    phone    = col(row, 11)
    visit_date  = col(row, 12)
    review      = col(row, 13)
    rating_taste = parse_rating(row, 14, rating)
    rating_value = parse_rating(row, 15, rating)
    rating_mood  = parse_rating(row, 16, rating)
    tip          = col(row, 17)
    tags_raw     = col(row, 18)
    photos_raw   = col(row, 19)
    map_url      = col(row, 20)

    cats = [c.strip() for c in cat_raw.split(',') if c.strip() in VALID_CATS]
    if not cats:
        cats = ['korean']

    tags   = [t.strip() for t in tags_raw.split(',')   if t.strip()] if tags_raw   else []
    photos = [p.strip() if '.' in p.strip() else p.strip() + '.jpg'
              for p in photos_raw.split(',') if p.strip()] if photos_raw else []

    return {
        'name': name, 'category': cats, 'description': desc,
        'location': location, 'menu': menu, 'rating': rating, 'image': image,
        'address': address, 'hours': hours, 'price': price,
        'parking': parking, 'phone': phone, 'visit_date': visit_date,
        'review': review, 'rating_taste': rating_taste,
        'rating_value': rating_value, 'rating_mood': rating_mood,
        'tip': tip, 'tags': tags, 'photos': photos, 'map_url': map_url,
    }


def read_xlsx():
    import openpyxl
    wb = openpyxl.load_workbook(XLSX_FILE)
    ws = wb.active
    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        parsed = parse_row(list(row))
        if parsed:
            results.append(parsed)
    return results


def create_xlsx_template():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '맛집데이터'

    header_fill = PatternFill('solid', fgColor='FFF3CD')
    bold_font   = Font(bold=True, size=10)
    center      = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wrap        = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font      = bold_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = thin_border
    ws.row_dimensions[1].height = 30

    for sample in SAMPLE_ROWS:
        ws.append(sample)

    # 행 높이 및 셀 서식
    for row_idx in range(2, len(SAMPLE_ROWS) + 2):
        ws.row_dimensions[row_idx].height = 60
        for cell in ws[row_idx]:
            cell.alignment = wrap
            cell.border    = thin_border

    # 열 너비
    col_widths = [22, 20, 36, 20, 28, 7, 14,
                  28, 28, 22, 22, 16,
                  14, 40, 8, 10, 10,
                  36, 30, 20, 28]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    # 카테고리 드롭다운 (B열)
    try:
        from openpyxl.worksheet.datavalidation import DataValidation
        dv = DataValidation(
            type='list',
            formula1='"korean,chinese,japanese,cafe,bar,lunch"',
            showDropDown=False
        )
        ws.add_data_validation(dv)
        dv.add('B2:B500')
    except Exception:
        pass

    ws.freeze_panes = 'A2'
    wb.save(XLSX_FILE)
    print('  ✅ 엑셀 템플릿 생성: ' + XLSX_FILE)


def read_csv():
    results = []
    with open(CSV_FILE, encoding='utf-8-sig', newline='') as f:
        reader = csv.reader(f)
        next(reader)
        for row in reader:
            while len(row) < len(HEADERS):
                row.append('')
            parsed = parse_row(row)
            if parsed:
                results.append(parsed)
    return results


def create_csv_template():
    with open(CSV_FILE, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        writer.writerows(SAMPLE_ROWS)
    print('  ✅ CSV 템플릿 생성: ' + CSV_FILE)
    print('     → Excel에서 열어 수정 후 CSV 형식으로 저장하세요.')


def write_json(restaurants):
    os.makedirs('data', exist_ok=True)
    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(restaurants, f, ensure_ascii=False, indent=2)
    print('  ✅ ' + str(len(restaurants)) + '개 맛집 → ' + OUTPUT_JSON + ' 저장 완료')


def main():
    print('=' * 45)
    print('  용문동 맛집 블로그 · 데이터 변환 도구')
    print('=' * 45)
    print()

    has_openpyxl = False
    try:
        import openpyxl
        has_openpyxl = True
    except ImportError:
        pass

    restaurants = None

    if os.path.exists(XLSX_FILE):
        if has_openpyxl:
            print('📂 ' + XLSX_FILE + ' 읽는 중...')
            restaurants = read_xlsx()
        else:
            print('⚠️  openpyxl이 없어 xlsx를 읽을 수 없습니다.')
            print('   아래 명령을 실행한 후 다시 시도하세요:')
            print('     pip install openpyxl')
            return

    if restaurants is None and os.path.exists(CSV_FILE):
        print('📂 ' + CSV_FILE + ' 읽는 중...')
        restaurants = read_csv()

    if restaurants is None:
        print('데이터 파일이 없습니다. 템플릿을 새로 만듭니다...\n')
        if has_openpyxl:
            create_xlsx_template()
            print('\n→ ' + XLSX_FILE + ' 을 Excel로 열어 맛집 정보를 입력하세요.')
        else:
            create_csv_template()
            print('\n→ ' + CSV_FILE + ' 을 Excel로 열어 맛집 정보를 입력하세요.')
        print('→ 저장 후 이 스크립트(데이터변환.bat)를 다시 실행하면 웹사이트가 업데이트됩니다.')
        return

    if not restaurants:
        print('⚠️  데이터가 비어 있습니다. 엑셀에 내용을 입력 후 다시 실행하세요.')
        return

    write_json(restaurants)
    print()
    print('✨ 완료! git commit 후 Vercel에서 자동 배포됩니다.')


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print('\n❌ 오류 발생: ' + str(e))
        import traceback
        traceback.print_exc()
    finally:
        print()
        input('엔터를 누르면 창이 닫힙니다...')
